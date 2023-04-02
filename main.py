from typing import Callable, Any
import os
import sys
from PyQt5 import QtCore, QtWidgets
from PyQt5.QtWidgets import QMessageBox
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Font, Side
import re
import datetime
import time
import json
import shapefile
import functools
from traceback import format_exc
from logic import write_settings, get_settings, to_shorten_a_long_name, extract_all_zipfiles
from real_estate import AbstractRealEstateObject
import graphic_interface

# делаем текущей директорией для работы ту папку, в которой лежит файл скрипта
path_to_current_file = os.path.realpath(__file__)
os.chdir(os.path.split(path_to_current_file)[0])

__author__ = "Dmitry S. Korottsev"
__copyright__ = "Copyright 2023"
__credits__ = []
__license__ = "GPL v3"
__version__ = "1.10"
__maintainer__ = "Dmitry S. Korottsev"
__email__ = "dm-korottev@yandex.ru"
__status__ = "Development"


def logger(func: Callable) -> Callable:
    """
    декоратор-логгер, записывает в файл "log.txt" ошибки, возникшие при работе функции, не изменяя имя исходной фукнкции
    """
    @functools.wraps(func)
    def wrapped(arg: Any) -> Any:
        try:
            result = func(arg)
            return result
        except:
            with open("log.txt", "a") as f:
                for s in format_exc().splitlines():
                    t = datetime.datetime.now()
                    f.write(t.strftime("%d.%m.%Y  %H:%M:%S") + " " + s + "\n")
                f.write("---------------------------------------------------------------------------------------------")
            print('Ошибка! Подробности в файле "log.txt"')
    return wrapped


class ConvXMLApp(QtWidgets.QMainWindow, graphic_interface.Ui_MainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setupUi(self)  # Это нужно для инициализации нашего дизайна
        # соединяем события с функциями
        self.btnBrowseIn.clicked.connect(self.browse_folder_in_xml)
        self.btnBrowseOut.clicked.connect(self.browse_folder_out_xml)
        self.btnStart.clicked.connect(self.start_conv)
        self.checkBoxShape.stateChanged.connect(self.change_check_box_shape)
        self.checkBoxExcel.stateChanged.connect(self.change_check_box_xlsx)
        self.checkBoxRename.stateChanged.connect(self.change_check_box_rename)
        self.checkBoxAdm.stateChanged.connect(self.change_check_box_adm)
        self.checkBoxReplace.stateChanged.connect(self.change_check_box_replace)
        #  используем ранее сохранённые настройки как значения по умолчанию
        with open('settings.json', 'r') as f:
            sd = json.load(f)
        self.label_input.setText(sd['folder_in_xml'])
        self.label_out.setText(sd['folder_out_xml'])

        if sd['create_esri_shape']:
            self.checkBoxShape.setCheckState(QtCore.Qt.Checked)
        else:
            self.checkBoxShape.setCheckState(QtCore.Qt.Unchecked)

        if sd['create_xlsx']:
            self.checkBoxExcel.setCheckState(QtCore.Qt.Checked)
        else:
            self.checkBoxExcel.setCheckState(QtCore.Qt.Unchecked)

        if sd['rename_files']:
            self.checkBoxRename.setCheckState(QtCore.Qt.Checked)
        else:
            self.checkBoxRename.setCheckState(QtCore.Qt.Unchecked)

        if sd['adm_district']:
            self.checkBoxAdm.setCheckState(QtCore.Qt.Checked)
        else:
            self.checkBoxAdm.setCheckState(QtCore.Qt.Unchecked)

        if sd['replace_long_names']:
            self.checkBoxReplace.setCheckState(QtCore.Qt.Checked)
        else:
            self.checkBoxReplace.setCheckState(QtCore.Qt.Unchecked)

    #  в случае изменения настроек записываем их в файл
    def change_check_box_shape(self) -> None:
        if self.checkBoxShape.isChecked():
            write_settings('create_esri_shape', True)
        else:
            write_settings('create_esri_shape', False)

    def change_check_box_xlsx(self) -> None:
        if self.checkBoxExcel.isChecked():
            write_settings('create_xlsx', True)
        else:
            write_settings('create_xlsx', False)

    def change_check_box_rename(self) -> None:
        if self.checkBoxRename.isChecked():
            write_settings('rename_files', True)
        else:
            write_settings('rename_files', False)

    def change_check_box_adm(self) -> None:
        if self.checkBoxAdm.isChecked():
            write_settings('adm_district', True)
        else:
            write_settings('adm_district', False)

    def change_check_box_replace(self) -> None:
        if self.checkBoxReplace.isChecked():
            write_settings('replace_long_names', True)
        else:
            write_settings('replace_long_names', False)

    def browse_folder_in_xml(self) -> None:
        directory = QtWidgets.QFileDialog.getExistingDirectory(self, "Выберите папку с выписками из ЕГРН в формате XML")
        # открыть диалог выбора директории и установить значение переменной
        # равной пути к выбранной директории
        if directory:  # не продолжать выполнение, если пользователь не выбрал директорию
            self.label_input.setText(str(directory))
            write_settings('folder_in_xml', str(directory))

    def browse_folder_out_xml(self) -> None:
        directory_out = QtWidgets.QFileDialog.getExistingDirectory(self, "Выберите папку, в которую нужно сохранить"
                                                                         "результат")
        if directory_out:
            self.label_out.setText(str(directory_out))
            write_settings('folder_out_xml', str(directory_out))

    def browse_folder_in_zip(self) -> None:
        directory_in_zip = QtWidgets.QFileDialog.getExistingDirectory(self,
                                                                      "Выберите папку с zip-архивами выписок из ЕГРН")
        if directory_in_zip:
            self.label_input_zip.setText(str(directory_in_zip))
            write_settings('folder_in_zip', str(directory_in_zip))

    @logger
    def extract_xml_from_zip(self) -> None:
        """
        извлекает выписки из ЕГРН из архива zip, сохраняя исходный архив, удаляя промежуточные
        архивы и файлы ЭЦП
        """
        self.textBrowser.append("Идёт извлечение выписок xml из архивов...")
        directory: str = get_settings('folder_in_xml')
        files = os.listdir(directory)
        zipfiles = list(filter(lambda x: x.endswith('.zip'), files))
        extract_all_zipfiles(zipfiles, directory)
        new_files = os.listdir(directory)
        new_zipfiles = list(filter(lambda x: x.endswith('.zip'), new_files))
        for i in new_zipfiles:
            if i in zipfiles:
                new_zipfiles.remove(i)
        extract_all_zipfiles(new_zipfiles, directory)
        result_files = os.listdir(directory)
        sig_files = list(filter(lambda x: x.endswith('.sig'), result_files))
        for sf in sig_files:
            os.remove(os.path.join(directory, sf))
        for zf in new_zipfiles:
            if zf not in zipfiles:
                os.remove(os.path.join(directory, zf))
        self.textBrowser.append("Извлечение выписок xml из архивов завершено.")
        self.textBrowser.append("------------------------------------------------------------------------------------"
                                "---------------------------")

    @logger
    def rename_xml(self) -> None:
        """
        переименовывает выписки из ЕГРН на земельные участки в формате: кадастровый номер---дата получения выписки
        """
        self.textBrowser.append("Идёт переименование выписок xml...")
        directory: str = get_settings('folder_in_xml')
        result_files = os.listdir(directory)
        pb = 0
        self.progressBar.setValue(0)
        count_successful_files = 0
        count_unsupported_files = 0
        xmlfiles = list(filter(lambda x: x.endswith('.xml'), result_files))
        for file_name in xmlfiles:
            xml_file_path = os.path.join(directory, file_name)
            real_estate_object = None
            if AbstractRealEstateObject.create_a_real_estate_object(xml_file_path):
                real_estate_object = AbstractRealEstateObject.create_a_real_estate_object(xml_file_path)
            if real_estate_object is not None:
                parcel_kn = real_estate_object.parent_cad_number
                extract_date = real_estate_object.extract_date
                pkn = re.sub(':', '-', parcel_kn)
                ed = re.sub('\.', '-', extract_date)
                new_name = pkn + '---' + ed + '.xml'
                new_path = os.path.join(directory, new_name)
                if file_name != new_name:
                    if not os.path.exists(new_path):
                        os.rename(os.path.join(directory, file_name), new_path)
                    else:
                        for num in range(2, 100):
                            new_path_double = os.path.join(directory, pkn + '---' + ed + ' (' + str(num) + ')' + '.xml')
                            if not os.path.exists(new_path_double):
                                os.rename(os.path.join(directory, file_name), new_path_double)
                                break
                    count_successful_files += 1
            else:
                count_unsupported_files += 1
            pb += 1
            self.progressBar.setValue(int((pb / len(xmlfiles)) * 100))
        files_do_not_require_renaming = len(xmlfiles) - count_successful_files - count_unsupported_files
        self.textBrowser.append("Готово!")
        if files_do_not_require_renaming > 0:
            self.textBrowser.append('Для ' + str(files_do_not_require_renaming) + ' xml-файлов переименование '
                                    'не требуется')
        self.textBrowser.append("Переименовано " + str(count_successful_files) + ' xml-файлов')
        if count_unsupported_files > 0:
            self.textBrowser.append("Не удалось прочитать " + str(count_unsupported_files) + ' xml-файлов')
        self.textBrowser.append("------------------------------------------------------------------------------------"
                                "---------------------------")

    @logger
    def start_conv(self):
        """
        запускает конвертирование набора выписок на земельные участки из формата xml в выбранные форматы файлов
        """
        directory = get_settings('folder_in_xml')
        starting_xmlfiles = list(filter(lambda x: x.endswith('.xml'), os.listdir(directory)))
        if self.radioButton_zip.isChecked() is False and self.radioButton_xml.isChecked() is False:
            QMessageBox.warning(self, 'Ошибка', "Необходимо выбрать формат обрабатываемых файлов (xml или zip)")
            return False
        elif self.radioButton_xml.isChecked() is True and len(starting_xmlfiles) == 0:
            QMessageBox.warning(self, 'Ошибка', "В указанной папке нет выписок из ЕГРН в формате XML")
            return False
        elif self.radioButton_zip.isChecked():
            self.extract_xml_from_zip()
        if self.checkBoxRename.isChecked():
            self.rename_xml()
        if self.checkBoxExcel.isChecked() or self.checkBoxShape.isChecked():
            xmlfiles = list(filter(lambda x: x.endswith('.xml'), os.listdir(directory)))
            self.textBrowser.append("Идёт получение данных из выписок XML и запись в выбранные форматы файлов...")
            start_time = time.time()
            now = datetime.datetime.now()
            directory_out = get_settings('folder_out_xml')
            if self.checkBoxExcel.isChecked():
                wb = Workbook()
                ws = wb.active
                fill_1 = PatternFill(fill_type='solid',
                                     start_color='c1c1c1',
                                     end_color='c2c2c2')
                font_1 = Font(name='Calibri',
                              size=11,
                              bold=True,
                              italic=False,
                              vertAlign=None,
                              underline='none',
                              strike=False,
                              color='FF000000')
                #  описание стиля границы ячейки в таблице MS Excel
                border_1 = Border(left=Side(border_style='thin',
                                            color='FF000000'),
                                  right=Side(border_style='thin',
                                             color='FF000000'),
                                  top=Side(border_style='thin',
                                           color='FF000000'),
                                  bottom=Side(border_style='thin',
                                              color='FF000000'),
                                  diagonal=Side(border_style='thin',
                                                color='FF000000'),
                                  diagonal_direction=0,
                                  outline=Side(border_style='thin',
                                               color='FF000000'),
                                  vertical=Side(border_style='thin',
                                                color='FF000000'),
                                  horizontal=Side(border_style='thin',
                                                  color='FF000000')
                                  )
                ws['A1'] = 'Кадастровый номер'
                ws['B1'] = 'Кадастровый номер единого землепользования'
                ws['C1'] = 'Площадь, м2'
                ws['D1'] = 'Адрес'
                ws['E1'] = 'Статус'
                ws['F1'] = 'Категория земель'
                ws['G1'] = 'Вид разрешенного использования (по документу)'
                ws['H1'] = 'Правообладатель'
                ws['I1'] = 'Вид права, номер и дата регистрации'
                ws['J1'] = 'Ограничения прав и обременения'
                ws['K1'] = 'Вид ограничения (обременения), номер и дата регистрации, срок действия'
                ws['L1'] = 'Особые отметки'
                ws['M1'] = 'Дата постановки на кад. учёт'
                ws['N1'] = 'Дата получения сведений'
                ws['O1'] = 'КН расположенных в пределах ЗУ или ОКС объектов недвижимости'
                ws['P1'] = 'Кадастровая стоимость, руб.'
                ws['Q1'] = 'Вид объекта недвижимости'
                for cell_obj in ws['A1':'Q1']:
                    for cell in cell_obj:
                        cell.fill = fill_1
                        cell.font = font_1
                ws.column_dimensions['A'].width = 18
                ws.column_dimensions['B'].width = 19
                ws.column_dimensions['C'].width = 10
                ws.column_dimensions['D'].width = 35
                ws.column_dimensions['E'].width = 16
                ws.column_dimensions['F'].width = 23
                ws.column_dimensions['G'].width = 37
                ws.column_dimensions['H'].width = 37
                ws.column_dimensions['I'].width = 45
                ws.column_dimensions['J'].width = 45
                ws.column_dimensions['K'].width = 45
                ws.column_dimensions['L'].width = 45
                ws.column_dimensions['M'].width = 14
                ws.column_dimensions['N'].width = 14
                ws.column_dimensions['O'].width = 18
                ws.column_dimensions['P'].width = 14
                ws.column_dimensions['Q'].width = 20
                row_numb = 1
            if self.checkBoxShape.isChecked():
                shp_wr = shapefile.Writer(os.path.join(directory_out, 'real_estate_objects_EGRN_' + now.strftime("%d_%m_%Y  %H-%M")),
                                          shapeType=shapefile.POLYGON, encoding="cp1251")
                shp_wr.field('CadNumber', 'C', size=20)
                shp_wr.field('SnglUseCN', 'C', size=20)
                shp_wr.field('NumOfCont', 'C', size=20)
                shp_wr.field('Area', 'N', 20, 2)
                shp_wr.field('Note', 'C', size=255)
                shp_wr.field('Parcel_St', 'C', size=255)
                shp_wr.field('Category', 'C', size=255)
                shp_wr.field('ByDoc', 'C', size=255)
                shp_wr.field('Owner', 'C', size=255)
                shp_wr.field('OwnRightN', 'C', size=255)
                shp_wr.field('Encumbr', 'C', size=255)
                shp_wr.field('EncRightN', 'C', size=255)
                shp_wr.field('Special', 'C', size=255)
                shp_wr.field('DatOfCreat', 'D')
                shp_wr.field('DateOfGet', 'D')
                shp_wr.field('EstateObjs', 'C', size=255)
                shp_wr.field('CadastCost', 'C', size=50)
                shp_wr.field('Type', 'C', size=60)
            xml_errors = []
            pb = 0
            count_successful_files = 0
            self.progressBar.setValue(0)
            for xml_file in xmlfiles:
                xml_file_path = os.path.join(directory, xml_file)
                real_estate_object = None
                if AbstractRealEstateObject.create_a_real_estate_object(xml_file_path):
                    real_estate_object = AbstractRealEstateObject.create_a_real_estate_object(xml_file_path)
                if real_estate_object is not None:
                    parent_cad_number = real_estate_object.parent_cad_number
                    entry_parcels = real_estate_object.entry_parcels
                    area = real_estate_object.area
                    #  с помощью регулярных выражений удаляем из строк символы табуляции, новой строки
                    #  и возврата каретки
                    pattern = "^\s+|\n|\r|\s+$"
                    address = re.sub(pattern, '', real_estate_object.address)
                    status = re.sub(pattern, '', real_estate_object.status)
                    category = real_estate_object.category
                    permitted_use_by_doc = re.sub(pattern, '', real_estate_object.permitted_use_by_doc)
                    owner = re.sub(pattern, '', real_estate_object.owner)
                    own_name_reg_numb_date = real_estate_object.own_name_reg_numb_date
                    encumbrances = re.sub(pattern, '', real_estate_object.encumbrances)
                    encumbrances_name_reg_numb_date_duration = real_estate_object.encumbrances_name_reg_numb_date_duration
                    special_notes = re.sub(pattern, '', real_estate_object.special_notes)
                    date_of_cadastral_reg = real_estate_object.date_of_cadastral_reg
                    extract_date = real_estate_object.extract_date
                    estate_objects = real_estate_object.estate_objects
                    cadastral_cost = real_estate_object.cadastral_cost
                    real_estate_object_type = real_estate_object.type
                    if self.checkBoxReplace.isChecked():
                        address = to_shorten_a_long_name(address)
                        permitted_use_by_doc = to_shorten_a_long_name(permitted_use_by_doc)
                        owner = to_shorten_a_long_name(owner)
                        encumbrances = to_shorten_a_long_name(encumbrances)
                        special_notes = to_shorten_a_long_name(special_notes)
                    if self.checkBoxShape.isChecked():
                        geometry = real_estate_object.geometry
                        if geometry != {}:
                            for key, value in geometry.items():
                                shp_wr.poly(value)
                                inverted_date_of_cadastral_reg = date_of_cadastral_reg.split(".")[::-1]
                                if inverted_date_of_cadastral_reg != ['']:
                                    year1, month1, day1 = inverted_date_of_cadastral_reg
                                else:
                                    year1, month1, day1 = 1, 1, 1
                                inverted_extract_date = extract_date.split(".")[::-1]
                                if inverted_extract_date != ['']:
                                    year2, month2, day2 = inverted_extract_date
                                else:
                                    year2, month2, day2 = 1, 1, 1
                                if re.search(r'\(', key):
                                    shp_cad_number = key[:key.index('(')]
                                    num_of_cont = key[key.index('('):]
                                elif not re.search(":", key):
                                    shp_cad_number = parent_cad_number
                                    num_of_cont = key
                                else:
                                    shp_cad_number = key
                                    num_of_cont = ''
                                if parent_cad_number == shp_cad_number:
                                    shp_parent_cad_number = ''
                                else:
                                    shp_parent_cad_number = parent_cad_number
                                shp_wr.record(shp_cad_number, shp_parent_cad_number, num_of_cont, float(area), address,
                                              status, category, permitted_use_by_doc, owner, own_name_reg_numb_date,
                                              encumbrances, encumbrances_name_reg_numb_date_duration, special_notes,
                                              datetime.date(int(year1), int(month1), int(day1)),
                                              datetime.date(int(year2), int(month2), int(day2)), estate_objects,
                                              cadastral_cost, real_estate_object_type)
                        else:
                            self.textBrowser.append(f'Выписка {xml_file} не содержит координат границ')
                    if self.checkBoxExcel.isChecked():
                        if not entry_parcels:
                            row_numb += 1
                            ws['A' + str(row_numb)] = parent_cad_number
                            ws['B' + str(row_numb)] = '-'
                            ws['C' + str(row_numb)] = area
                            ws['D' + str(row_numb)] = address
                            ws['E' + str(row_numb)] = status
                            ws['F' + str(row_numb)] = category
                            ws['G' + str(row_numb)] = permitted_use_by_doc
                            ws['H' + str(row_numb)] = owner
                            ws['I' + str(row_numb)] = own_name_reg_numb_date
                            ws['J' + str(row_numb)] = encumbrances
                            ws['K' + str(row_numb)] = encumbrances_name_reg_numb_date_duration
                            ws['L' + str(row_numb)] = special_notes
                            ws['M' + str(row_numb)] = date_of_cadastral_reg
                            ws['N' + str(row_numb)] = extract_date
                            ws['O' + str(row_numb)] = estate_objects
                            ws['P' + str(row_numb)] = cadastral_cost
                            ws['Q' + str(row_numb)] = real_estate_object_type
                        else:
                            for parcel_cad_number in entry_parcels:
                                row_numb += 1
                                ws['A' + str(row_numb)] = parcel_cad_number
                                ws['B' + str(row_numb)] = parent_cad_number
                                ws['C' + str(row_numb)] = area
                                ws['D' + str(row_numb)] = address
                                ws['E' + str(row_numb)] = status
                                ws['F' + str(row_numb)] = category
                                ws['G' + str(row_numb)] = permitted_use_by_doc
                                ws['H' + str(row_numb)] = owner
                                ws['I' + str(row_numb)] = own_name_reg_numb_date
                                ws['J' + str(row_numb)] = encumbrances
                                ws['K' + str(row_numb)] = encumbrances_name_reg_numb_date_duration
                                ws['L' + str(row_numb)] = special_notes
                                ws['M' + str(row_numb)] = date_of_cadastral_reg
                                ws['N' + str(row_numb)] = extract_date
                                ws['O' + str(row_numb)] = estate_objects
                                ws['P' + str(row_numb)] = cadastral_cost
                                ws['Q' + str(row_numb)] = real_estate_object_type
                    if self.checkBoxShape.isChecked():
                        pass
                    count_successful_files += 1
                else:
                    xml_errors.append(xml_file_path)
                pb += 1
                self.progressBar.setValue(int((pb / len(xmlfiles)) * 100))
            if self.checkBoxExcel.isChecked():
                for cell_obj in ws['A1':'Q' + str(row_numb)]:
                    for cell in cell_obj:
                        cell.border = border_1
                        cell.alignment = Alignment(wrapText=True)  # задаёт выравнивание "перенос по словам"
                wb.save(os.path.join(directory_out, now.strftime("%d_%m_%Y  %H-%M") +
                        " real_estate_objects_EGRN.xlsx"))
            if self.checkBoxShape.isChecked():
                shp_wr.close()
            self.textBrowser.append("Получение данных из выписок XML завершено!" + chr(13) +
                                    "Результат сохранён в папке " + directory_out)
            sec = round(float("%s" % (time.time() - start_time)))
            if sec == 0:
                sec = 1
            self.textBrowser.append("Успешно обработано " + str(count_successful_files) + " файлов за " + str(sec) +
                                    " сек.")
            if len(xml_errors) > 0:
                self.textBrowser.append("Не обработано " + str(len(xml_errors)) + " файлов:")
                for err_file in xml_errors:
                    self.textBrowser.append(err_file)
            self.textBrowser.append("----------------------------------------------------------------------------------"
                                    "-----------------------------")


def main():
    if not os.path.exists('settings.json'):
        s_dic = {'folder_in_xml': '', 'folder_out_xml': '', 'file_type': 'xml', 'create_esri_shape': False,
                 'create_xlsx': True, 'rename_files': True, 'adm_district': False, 'replace_long_names': True}
        with open('settings.json', 'w') as f:
            json.dump(s_dic, f, sort_keys=True, indent=4, ensure_ascii=False)

    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    app = QtWidgets.QApplication(sys.argv)  # новый экземпляр QApplication
    app.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)  # автоматически адаптирует интерфейс для 4K монитора
    window = ConvXMLApp()  # создаём объект класса ConvXMLApp
    window.show()  # показываем окно
    window.setFixedSize(559, 852)  # устанавливаем фиксированный размер окна
    app.exec_()  # запускаем приложение


if __name__ == "__main__":
    main()
